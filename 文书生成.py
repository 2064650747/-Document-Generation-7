import os  # 导入操作系统模块
import warnings  # 导入警告模块
import pandas as pd  # 导入pandas模块并简写为pd
from datetime import datetime  # 从datetime模块导入datetime类
from docx import Document  # 从docx模块导入Document类
from openpyxl import load_workbook  # 从openpyxl模块导入load_workbook函数
from openpyxl.styles import Alignment, Border, Side, Font  # 从openpyxl.styles模块导入样式类
from openpyxl.drawing.image import Image  # 从openpyxl.drawing模块导入Image类
import pdfplumber  # 导入pdfplumber模块
import poexcel  # 导入poexcel模块
import poword  # 导入poword模块
import re  # 导入正则表达式模块
from loguru import logger  # 从loguru模块导入logger

warnings.filterwarnings('ignore')  # 忽略所有警告

# 定义样式
alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
font = Font(size=14)

# 图片路径
image_paths = [r"D:\工作案件\1.12凌源专案\凌源市文书生成\凌源公章.png"]

# 定义文件夹路径
folder_path = r'F:\凌源市文书生成\结果\测试'

# 指定要插入图片的单元格位置列表
cell_positions = [
    "C6", "C47", "C89", "C132", "C173", "C215", "C257", "C299", "C341", "C383",
    "C425", "C467", "C509", "C551", "C593", "C635", "C677", "C719", "C761",
    "C803", "C845", "C887", "C929", "C971"
]

def listdir(path, ln):  # 定义函数listdir，用于获取所有文件路径
    for file in os.listdir(path):  # 遍历目录中的每个文件
        file_path = os.path.join(path, file)  # 获取文件的完整路径
        if os.path.isdir(file_path):  # 如果是目录，递归调用listdir
            listdir(file_path, ln)
        else:  # 如果是文件，添加到列表ln
            ln.append(file_path)
    return ln  # 返回文件路径列表

def text_create(target: dict):  # 定义函数text_create，用于创建文本描述
    names = list(target.keys())  # 获取字典的所有键
    text_1 = '、'.join(names)  # 将键用顿号连接成字符串
    text_list = []  # 初始化文本列表
    for n in names:  # 遍历每个键
        res = '{0}{1}等{2}张银行卡'.format(n, target[n][0], target[n][1])  # 格式化字符串
        text_list.append(res)  # 添加到文本列表
    text_2 = '、'.join(text_list)  # 将文本列表用顿号连接成字符串
    return text_1, text_2  # 返回两个文本字符串

def get_word_new(ln, moban, num, save_path):  # 定义函数get_word_new，用于生成Word和Excel文件
    redate = datetime.now()  # 获取当前日期时间
    day = '{}年{}月{}日'.format(redate.year, redate.month, redate.day)  # 格式化日期字符串
    for file in ln:  # 遍历文件路径列表
        print(f'读取文件"{file}"')  # 打印读取文件信息
        (filepath, filename) = os.path.split(file)  # 分割文件路径和文件名
        file_save_path = os.path.join(save_path, filename.split('.')[0])  # 生成保存路径
        if not os.path.exists(file_save_path):  # 如果保存路径不存在，创建目录
            os.makedirs(file_save_path)
        tup = pd.read_excel(file, dtype=object)  # 读取Excel文件
        start = 0  # 初始化处理位置
        total_rows = len(tup)  # 获取总行数
        while start < total_rows:  # 当未处理完所有行时
            banks_list = []  # 初始化银行列表
            batch_tup = pd.DataFrame()  # 初始化数据批次
            while len(banks_list) < 7 and len(batch_tup) < 1000 and start < total_rows:  # 收集不超过7个银行和不超过1000行的数据
                row = tup.iloc[start]  # 获取当前行
                bank = row['查询账户所属银行']  # 获取银行名称
                if bank not in banks_list:  # 如果银行不在列表中
                    if len(banks_list) < 7:  # 如果银行列表少于7个，添加银行
                        banks_list.append(bank)
                batch_tup = pd.concat([batch_tup, row.to_frame().T], ignore_index=True)
                start += 1  # 处理下一行
            bank_dict = {}  # 创建银行字典
            for bank in banks_list:  # 遍历银行列表
                filtered = batch_tup[batch_tup['查询账户所属银行'] == bank]['查询账(卡)号'].iloc[0]  # 获取第一个账号
                nums = batch_tup.loc[batch_tup['查询账户所属银行'] == bank]['查询账(卡)号'].count()  # 计算账号数量
                bank_dict[bank] = (filtered, nums)  # 添加到银行字典
            t1, t2 = text_create(bank_dict)  # 创建文本描述
            docStr = Document(moban)  # 打开Word模板
            children = docStr.element.body.iter()  # 获取文档元素
            for child in children:  # 遍历文档元素
                for ci in child.iter():  # 遍历子元素
                    if ci.tag.endswith('main}r'):  # 如果是文本元素
                        if 'n' in ci.text:  # 替换占位符n1
                            ci.text = ci.text.replace('n1', str(num))
                        if 'z' in ci.text:  # 替换占位符z
                            ci.text = ci.text.replace('z', str(num))
                        if 'm' in ci.text:  # 替换占位符m
                            ci.text = ci.text.replace('m', t2)
                        if 'w' in ci.text:  # 替换占位符w
                            ci.text = ci.text.replace('w', t1)
                        if 'd' in ci.text:  # 替换占位符d
                            ci.text = ci.text.replace('d', day)
            filtered_tup = (
                batch_tup[batch_tup['查询账户所属银行'].isin(bank_dict.keys())]
                .assign(
                    账卡号类型='个人',
                    查询种类='账户及交易明细',
                    时间标识='自定义时间段',
                    开始时间='2022-01-01',  # 修改至具体日期
                    结束时间='2024-11-06'   # 修改至具体日期
                      )
                .rename(columns={'查询账(卡)号': '被查账/卡号', '查询账户所属银行': '选择银行'})
                .sort_values(by='选择银行', ascending=False)
                .reset_index(drop=True)
                [['账卡号类型', '被查账/卡号', '选择银行', '查询种类', '时间标识', '开始时间', '结束时间']])
            save_doc = os.path.join(file_save_path, f'协助查询财产通知书-{num}.docx')  # 生成Word保存路径
            save_xl = os.path.join(file_save_path, f'{num}.xlsx')  # 生成Excel保存路径
            docStr.save(save_doc)  # 保存Word文档
            filtered_tup.to_excel(save_xl, index=False)  # 保存Excel表格
            print(f"已生成文档: {save_doc}")  # 打印生成文档信息
            print(f"已生成Excel表格: {save_xl}")  # 打印生成Excel信息
            num += 1  # 增加编号

if __name__ == "__main__":  # 主程序入口
    模式 = '不加附表'  # 设置模式
    模板 = r"D:\工作案件\协助查询财产通知书.docx"  # 设置模板路径
    表格 = r"F:\凌源市文书生成\表格"  # 设置表格路径
    保存位置 = r"F:\凌源市文书生成\结果"  # 设置保存位置
    起始编号 = 2811  # 设置起始编号
    list_name = []  # 初始化文件列表
    list_name = listdir(表格, list_name)  # 获取表格文件列表
    get_word_new(list_name, 模板, 起始编号, 保存位置)  # 生成文档和表格

# 遍历文件夹中的所有文件
for file in os.listdir(folder_path):
    if file.endswith('.xls') or file.endswith('.xlsx'):
        file_path = os.path.join(folder_path, file)
        # 读取 Excel 文件
        df = pd.read_excel(file_path)
        # 选择需要的两列并转换数据类型
        df = df[['被查账/卡号', '选择银行']].astype({'被查账/卡号': str})
        df['序号'] = range(1, len(df) + 1)
        df = df.reindex(columns=['序号', '被查账/卡号', '选择银行'])
        # 生成新文件名
        if file.endswith('.xlsx'):
            new_file_name = file.replace('.xlsx', '_附件.xlsx')
        elif file.endswith('.xls'):
            new_file_name = file.replace('.xls', '_附件.xls')
        new_file_path = os.path.join(folder_path, new_file_name)

        # 创建 Excel writer 对象
        with pd.ExcelWriter(new_file_path, engine='xlsxwriter') as writer:
            # 将 DataFrame 写入 Excel
            df.to_excel(writer, index=False)

        if new_file_name.endswith('附件.xlsx'):
            try:
                # 加载工作簿
                wb = load_workbook(filename=new_file_path)

                # 遍历每个工作表
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    # 遍历每个单元格
                    for row in sheet.iter_rows():
                        for cell in row:
                            # 设置样式
                            cell.alignment = alignment
                            cell.border = thin_border
                            cell.font = font
                    # 调整第一列宽度为 8
                    sheet.column_dimensions['A'].width = 8

                    # 调整第二列宽度为 44
                    sheet.column_dimensions['B'].width = 27

                    # 调整第三列宽度为 20
                    sheet.column_dimensions['C'].width = 44

                    # 设置行高为 18
                    for row in sheet.rows:
                        sheet.row_dimensions[row[0].row].height = 18

                # 插入图片到确定的位置
                max_row = sheet.max_row
                for i, pos in enumerate(cell_positions):
                    cell_row = int(pos[1:])
                    if max_row >= cell_row:
                        image_path = image_paths[i % len(image_paths)]
                        img = Image(image_path)
                        img.anchor = pos
                        sheet.add_image(img, pos)

                # 保存工作簿
                wb.save(new_file_path)
                print(f"文件 '{new_file_name}' 处理成功。")

            except Exception as e:
                print(f"处理文件 '{new_file_name}' 时出现错误：{e}")

    if file.endswith(".xlsx") and not file.endswith("_附件.xlsx"):
        filepath = os.path.join(folder_path, file)
        df = pd.read_excel(filepath)
        str_columns = ["被查账/卡号"]
        for col in str_columns:
            df[col] = df[col].apply(str)
        new_filename = os.path.splitext(file)[0] + ".xls"
        new_filepath = os.path.join(folder_path, new_filename)
        df.to_excel(new_filepath, index=False, engine='xlsxwriter')

# 删除结尾是.xlsx 但不是 _附件.xlsx 的文件
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx") and not filename.endswith("_附件.xlsx"):
        file_to_delete = os.path.join(folder_path, filename)
        os.remove(file_to_delete)

for root, dirs, files in os.walk(folder_path):
    for file in files:
        file_path = os.path.join(root, file)
        logger.info(f"PDF 转换:{file_path}")
        if '附件.xlsx' in file:
            poexcel.excel2pdf(excel_path=file_path, pdf_path=f"{root}\JPG")
        if '.docx' in file:
            poword.docx2pdf(path=file_path, output_path=f"{root}\JPG")
logger.info(f"PDF 转换完毕")

# 保存为 JPG
for root, dirs, files in os.walk(folder_path):
    for file in files:
        if '.pdf' in file:
            file_path = os.path.join(root, file)
            logger.info(f"JPG 转换:{file_path}")
            pdf = pdfplumber.open(file_path)
            for index, page in enumerate(pdf.pages):
                logger.info(page)
                to_JPG = page.to_image(resolution=230)
                to_JPG.save(f"{root}\\{str(re.sub('.pdf', f'-{index + 1}.jpg', file))}")
logger.info(f"JPG 转换完毕")